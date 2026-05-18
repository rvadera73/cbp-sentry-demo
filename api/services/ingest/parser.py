"""
Manifest Excel parser — extracts rows from password-protected Excel files
"""
import logging
from typing import List, Dict, Any
import re
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def parse_excel_manifest(file_path: str, password: str = None) -> List[Dict[str, Any]]:
    """
    Parse Excel manifest file and extract rows.

    Args:
        file_path: Path to .xlsx file
        password: Optional password for protected sheet

    Returns:
        List of row dictionaries with keys: shipper, consignee, hts_code, quantity, value, description

    Raises:
        ValueError: If file is invalid or required columns missing
        IOError: If file cannot be read
    """
    try:
        # Load workbook (password parameter not reliably supported in all openpyxl versions)
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Use first sheet
        ws = wb.active

        if not ws:
            raise ValueError("Excel file has no sheets")

        rows = _extract_rows(ws)
        logger.info(f"Parsed {len(rows)} rows from {file_path}")

        return rows

    except Exception as e:
        logger.error(f"Failed to parse Excel manifest: {e}")
        raise


def _extract_rows(ws: Worksheet) -> List[Dict[str, Any]]:
    """Extract rows from worksheet, handling headers and normalization"""

    # Read header row (row 1)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(_normalize_header(cell.value))

    if not headers:
        raise ValueError("No headers found in Excel file")

    # Validate required columns (support both "quantity" and "quantity_kg", "value" and "value_usd")
    required_base = {"shipper", "consignee", "hts_code", "description"}
    found_columns = set(headers)

    # Check for quantity column (accept either "quantity" or "quantity_kg" or variants)
    quantity_col = next((h for h in headers if h.startswith("quantity")), None)
    if not quantity_col:
        raise ValueError("Missing required column: quantity")

    # Check for value column (accept either "value" or "value_usd" or variants)
    value_col = next((h for h in headers if h.startswith("value")), None)
    if not value_col:
        raise ValueError("Missing required column: value")

    # Check for other required columns
    missing = required_base - found_columns
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    # Map column indices
    col_indices = {
        "shipper": headers.index("shipper"),
        "consignee": headers.index("consignee"),
        "hts_code": headers.index("hts_code"),
        "quantity": headers.index(quantity_col),
        "value": headers.index(value_col),
        "description": headers.index("description"),
    }

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        cells = list(ws[row_idx])

        # Skip empty rows
        if all(cell.value is None for cell in cells):
            continue

        try:
            row_dict = {
                "shipper": _normalize_string(cells[col_indices["shipper"]].value),
                "consignee": _normalize_string(cells[col_indices["consignee"]].value),
                "hts_code": _normalize_hts_code(cells[col_indices["hts_code"]].value),
                "quantity_kg": _parse_number(cells[col_indices["quantity"]].value),
                "value_usd": _parse_number(cells[col_indices["value"]].value),
                "description": _normalize_string(cells[col_indices["description"]].value),
            }

            # Validate row
            if not all(row_dict.values()):
                logger.warning(f"Skipping row {row_idx}: missing required field")
                continue

            rows.append(row_dict)

        except (ValueError, TypeError) as e:
            logger.warning(f"Skipping row {row_idx}: {e}")
            continue

    return rows


def _normalize_header(header: str) -> str:
    """Normalize column header to lowercase underscore format"""
    header = str(header).strip().lower()
    header = re.sub(r'[\s\-\.]+', '_', header)
    header = re.sub(r'[^a-z0-9_]', '', header)
    return header


def _normalize_string(value: Any) -> str:
    """Normalize string: trim, deduplicate spaces"""
    if value is None:
        return ""

    s = str(value).strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def _normalize_hts_code(value: Any) -> str:
    """Normalize HTS code: remove spaces, validate format"""
    if value is None:
        raise ValueError("HTS code cannot be empty")

    code = str(value).strip().replace(" ", "").replace(".", "")

    # Validate: should be 10 digits
    if not re.match(r'^\d{10}$', code):
        raise ValueError(f"Invalid HTS code format: {code}")

    # Format as XXXX.XX.XXXX
    return f"{code[:4]}.{code[4:6]}.{code[6:10]}"


def _parse_number(value: Any) -> float:
    """Parse numeric value, handling currency symbols and formats"""
    if value is None:
        raise ValueError("Numeric value cannot be empty")

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    # Remove currency symbols and spaces
    s = re.sub(r'[$€£¥₹\s,]', '', s)

    try:
        return float(s)
    except ValueError:
        raise ValueError(f"Cannot parse number: {value}")
